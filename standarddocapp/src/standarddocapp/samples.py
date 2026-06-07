"""Programmatic sample-Excel generators (3GPP / IEEE / Search).

The harvest workbook builder accepts user-supplied OutputRootFolder /
JobName / ProxyURL plus an optional list of URLs (otherwise the built-in
defaults below are used).

Public API:

* :func:`build_harvest_sample` - generic, configurable builder.
* :func:`build_3gpp_sample` / :func:`build_ieee_sample` - thin wrappers used
  by the existing one-click About-tab buttons.
* :func:`build_search_sample` - the search workbook (unchanged).
* :func:`read_links_from_excel` - read column A of an arbitrary "links"
  workbook so users can pre-populate the C column of the produced
  sample_download.xlsx.
* :func:`default_output_root` - the user's Downloads folder.
"""
from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Tuple
from urllib.parse import unquote, urlsplit

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------- styles

_HEADER_FILL = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")
_LABEL_FILL = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
_LINK_FONT = Font(color="FF0563C1", underline="single")
_BOLD = Font(bold=True)


# ---------------------------------------------------------------- defaults

DEFAULT_3GPP_URLS: Tuple[str, ...] = (
    "https://www.3gpp.org/ftp/tsg_ran/WG1_RL1/TSGR1_120b/Docs/R1-2502624.zip",
    "https://www.3gpp.org/ftp/tsg_ran/WG1_RL1/TSGR1_120b/Docs/R1-2502715.zip",
    "https://www.3gpp.org/ftp/tsg_ran/WG1_RL1/TSGR1_120b/Docs/R1-2502726.zip",
    "https://www.3gpp.org/ftp/tsg_ran/WG1_RL1/TSGR1_120b/Docs/R1-2502776.zip",
    "https://www.3gpp.org/ftp/tsg_ran/WG1_RL1/TSGR1_120b/Docs/R1-2502814.zip",
    "https://www.3gpp.org/ftp/tsg_ran/WG1_RL1/TSGR1_120b/Docs/R1-2502821.zip",
)

DEFAULT_IEEE_URLS: Tuple[str, ...] = (
    "https://mentor.ieee.org/802.11/dcn/16/11-16-1424-02-00ax-issue-on-the-cwmax-value-in-mu-edca.pptx",
    "https://mentor.ieee.org/802.11/dcn/16/11-16-1425-02-00ax-clarification-on-applying-mu-edca-parameter-set.pptx",
    "https://mentor.ieee.org/802.11/dcn/16/11-16-1368-02-00ax-follow-up-on-mu-edca-parameters.docx",
    "https://mentor.ieee.org/802.11/dcn/16/11-16-0998-02-00ax-rules-for-2-edca-parameters.pptx",
    "https://mentor.ieee.org/802.11/dcn/16/11-16-0963-02-00ax-edca-rules-8211-follow-up-2.pptx",
    "https://mentor.ieee.org/802.11/dcn/16/11-16-0962-02-00ax-edca-rules-8211-follow-up-1.pptx",
)

DEFAULT_JOB_NAMES = {"3gpp": "3gpp_sample_job", "ieee": "ieee_sample_job"}


def default_output_root() -> Path:
    """Return the user's Downloads folder, falling back to the home directory."""
    if os.name == "nt":
        userprofile = os.environ.get("USERPROFILE")
        if userprofile:
            cand = Path(userprofile) / "Downloads"
            if cand.exists():
                return cand
    cand = Path.home() / "Downloads"
    return cand if cand.exists() else Path.home()


_FILENAME_RE = re.compile(r"[^A-Za-z0-9._-]+")


def _title_from_url(url: str) -> str:
    """Best-effort title derived from a URL (file name without extension)."""
    if not url:
        return ""
    try:
        path = urlsplit(url).path
        name = unquote(os.path.basename(path))
    except Exception:
        return url
    if not name:
        return url
    stem = name.rsplit(".", 1)[0] if "." in name else name
    if not stem:
        stem = name
    return stem


# ---------------------------------------------------------------- harvest

_SETTINGS_LABELS: Tuple[str, ...] = (
    "ProxyURL",
    "TimeoutSec",
    "RetryCount",
    "SleepSec",
    "OverwriteExisting",
    "MinFileSizeKB",
    "MaxFileSizeMB",
    "OnTooSmallFile",
    "OnTooLargeFile",
    "KillOfficeAppsBeforeRun",
    "DownloadWorkers",
    "UnzipWorkers",
    "PdfWorkers",
    "HtmlWorkers",
    "CombineHtmlBatchSize",
    # Optional proxy authentication (read by stdharvest from rows 16/17).
    "ProxyUser",
    "ProxyPassword",
)

_DEFAULT_SETTINGS_VALUES = {
    "ProxyURL": "",
    "TimeoutSec": 60,
    "RetryCount": 5,
    "SleepSec": 0.5,
    "OverwriteExisting": "no",
    "MinFileSizeKB": 10,
    "MaxFileSizeMB": 100,
    "OnTooSmallFile": "error",
    "OnTooLargeFile": "skip",
    "KillOfficeAppsBeforeRun": "yes",
    "DownloadWorkers": 8,
    "UnzipWorkers": 4,
    "PdfWorkers": 2,
    "HtmlWorkers": 6,
    "CombineHtmlBatchSize": 5,
    "ProxyUser": "",
    "ProxyPassword": "",
}

_SETTINGS_VALIDATIONS: Tuple[Tuple[str, str], ...] = (
    ("B5", '"yes,no"'),
    ("B8", '"error,skip"'),
    ("B9", '"skip,pdf_only,keep_raw"'),
    ("B10", '"yes,no"'),
)

_SHEET1_HEADERS = ["", "", "Link", "Status", "SavedPath", "Message", "LastRunAt"]


def _populate_harvest_sheet1(
    ws,
    *,
    source_type: str,
    output_root: str,
    job_name: str,
    rows: Iterable[Tuple[str, str]],
) -> None:
    ws.title = "Sheet1"

    pairs = (
        ("A1", "SourceType", "B1", source_type),
        ("A2", "OutputRootFolder", "B2", output_root),
        ("A3", "JobName", "B3", job_name),
    )
    for label_addr, label_text, value_addr, value_text in pairs:
        c = ws[label_addr]
        c.value = label_text
        c.font = _BOLD
        c.fill = _LABEL_FILL
        ws[value_addr] = value_text

    for col_idx, text in enumerate(_SHEET1_HEADERS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=text)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left")

    for i, (title, url) in enumerate(rows, start=5):
        cell = ws.cell(row=i, column=3, value=title)
        if url:
            try:
                cell.hyperlink = url
                cell.font = _LINK_FONT
            except Exception:
                pass

    widths = {"A": 22, "B": 40, "C": 56, "D": 14, "E": 40, "F": 40, "G": 22}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    dv_source = DataValidation(type="list", formula1='"3gpp,ieee"', allow_blank=False)
    dv_source.error = "3gpp か ieee を選択してください"
    dv_source.errorTitle = "SourceType"
    ws.add_data_validation(dv_source)
    dv_source.add("B1")


def _populate_settings_sheet(
    ws, *, proxy_url: str = "", proxy_user: str = "", proxy_password: str = "",
) -> None:
    overrides = {
        "ProxyURL": proxy_url,
        "ProxyUser": proxy_user,
        "ProxyPassword": proxy_password,
    }
    for i, label in enumerate(_SETTINGS_LABELS, start=1):
        value = overrides.get(label, _DEFAULT_SETTINGS_VALUES[label])
        c_a = ws.cell(row=i, column=1, value=label)
        c_a.font = _BOLD
        c_a.fill = _LABEL_FILL
        ws.cell(row=i, column=2, value=value)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 40

    for cell_ref, formula in _SETTINGS_VALIDATIONS:
        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(cell_ref)


def _resolve_url_rows(
    source_type: str,
    urls: Optional[Sequence[Tuple[str, str]] | Sequence[str]] = None,
) -> List[Tuple[str, str]]:
    """Normalise a URL list (or None) to ``[(title, url), ...]``.

    ``urls`` may be ``None`` (defaults), a list of URL strings, or a list of
    ``(title, url)`` pairs.  Strings are auto-titled from the URL filename.
    """
    if urls is None:
        if source_type == "3gpp":
            base = DEFAULT_3GPP_URLS
        elif source_type == "ieee":
            base = DEFAULT_IEEE_URLS
        else:
            base = ()
        return [(_title_from_url(u), u) for u in base]

    out: List[Tuple[str, str]] = []
    for entry in urls:
        if isinstance(entry, tuple):
            title, url = entry
        else:
            title = ""
            url = str(entry)
        url = (url or "").strip()
        title = (title or "").strip() or _title_from_url(url)
        if not url:
            continue
        out.append((title, url))
    return out


def build_harvest_sample(
    out_path: Path,
    *,
    source_type: str,
    output_root: Optional[str] = None,
    job_name: Optional[str] = None,
    proxy_url: str = "",
    proxy_user: str = "",
    proxy_password: str = "",
    urls: Optional[Sequence[Tuple[str, str]] | Sequence[str]] = None,
) -> Path:
    """Build a sample download.xlsx with full control over every Sheet1 field.

    Any value left as ``None`` falls back to the established sample default
    (Downloads folder for ``output_root``, ``<source>_sample_job`` for
    ``job_name``, the canonical IEEE/3GPP URL set for ``urls``).
    """
    src = (source_type or "").strip().lower()
    if src not in ("3gpp", "ieee"):
        raise ValueError(f"source_type must be '3gpp' or 'ieee' (got {source_type!r})")

    if not output_root:
        output_root = str(default_output_root())
    if not job_name:
        job_name = DEFAULT_JOB_NAMES[src]

    rows = _resolve_url_rows(src, urls)

    wb = Workbook()
    _populate_harvest_sheet1(
        wb.active,
        source_type=src,
        output_root=output_root,
        job_name=job_name,
        rows=rows,
    )
    _populate_settings_sheet(
        wb.create_sheet("Sheet2"),
        proxy_url=proxy_url or "",
        proxy_user=proxy_user or "",
        proxy_password=proxy_password or "",
    )

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def build_3gpp_sample(
    out_path: Path,
    *,
    output_root: Optional[str] = None,
    job_name: Optional[str] = None,
    proxy_url: str = "",
    proxy_user: str = "",
    proxy_password: str = "",
    urls: Optional[Sequence[Tuple[str, str]] | Sequence[str]] = None,
) -> Path:
    return build_harvest_sample(
        out_path, source_type="3gpp",
        output_root=output_root, job_name=job_name,
        proxy_url=proxy_url, proxy_user=proxy_user, proxy_password=proxy_password,
        urls=urls,
    )


def build_ieee_sample(
    out_path: Path,
    *,
    output_root: Optional[str] = None,
    job_name: Optional[str] = None,
    proxy_url: str = "",
    proxy_user: str = "",
    proxy_password: str = "",
    urls: Optional[Sequence[Tuple[str, str]] | Sequence[str]] = None,
) -> Path:
    return build_harvest_sample(
        out_path, source_type="ieee",
        output_root=output_root, job_name=job_name,
        proxy_url=proxy_url, proxy_user=proxy_user, proxy_password=proxy_password,
        urls=urls,
    )


# ---------------------------------------------------------------- read links

def read_links_from_excel(path: Path) -> List[Tuple[str, str]]:
    """Return ``[(title, url), ...]`` from column A of an arbitrary workbook.

    For each non-empty cell in column A starting at A1:

    * If the cell carries a hyperlink, ``cell.hyperlink.target`` becomes the URL
      and the cell's text becomes the title.
    * Otherwise the cell value is treated as the URL string and the title is
      derived from the URL.

    Cells without a usable URL are skipped silently.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(path)
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb.active
        out: List[Tuple[str, str]] = []
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=1)
            text = "" if cell.value is None else str(cell.value).strip()
            url = ""
            if cell.hyperlink is not None:
                target = getattr(cell.hyperlink, "target", "") or ""
                if target:
                    url = str(target).strip()
            if not url and text:
                if text.lower().startswith(("http://", "https://", "ftp://")):
                    url = text
            if not url:
                continue
            title = text or _title_from_url(url)
            out.append((title, url))
        return out
    finally:
        wb.close()


# ---------------------------------------------------------------- search

_SEARCH_COLUMNS: Tuple[Tuple[str, int], ...] = (
    ("Use", 7),
    ("RuleName", 22),
    ("MatchType", 12),
    ("Term1", 22),
    ("Term2", 22),
    ("Term3", 18),
    ("Term4", 18),
    ("Scope", 14),
    ("Notes", 28),
)

_SAMPLE_SEARCH_RULES: Tuple[Tuple[str, str, str, str, str, str, str, str, str], ...] = (
    ("yes", "bssid_single",      "SINGLE", "Multiple BSSID", "",                  "", "", "sentence",  "一語検索 (IEEE)"),
    ("yes", "bssid_sta_profile", "AND",    "Multiple BSSID", "STA profile",       "", "", "sentence",  "同一文にBSSIDとSTA profile"),
    ("yes", "mmwave_bssid",      "AND",    "mmWave",         "M-BSSID",           "", "", "block",     "スライド全体でAND (IEEE)"),
    ("yes", "dci_priority",      "AND",    "DCI format 0_3", "Priority indicator","", "", "paragraph", "同一段落AND (3GPP Word)"),
    ("no",  "example_disabled",  "SINGLE", "example",        "",                  "", "", "sentence",  "Use=no で無効化例"),
)


def build_search_sample(out_path: Path, *, project_name: str = "bssid_search",
                        job_folder: str = "", output_folder: str = "") -> Path:
    """Write a sample search.xlsx with a few canned rules."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "ProjectName"
    ws["B1"] = project_name
    ws["A2"] = "JobFolder"
    ws["B2"] = job_folder
    ws["A3"] = "OutputFolder"
    ws["B3"] = output_folder
    ws["A4"] = "(空欄可: JobFolder/search/YYYYMMDD_<ProjectName> に出力)"
    for row in (1, 2, 3):
        ws.cell(row=row, column=1).font = _BOLD
    note_fill = PatternFill("solid", fgColor="FFF7D6")
    ws.cell(row=4, column=1).fill = note_fill
    ws.cell(row=4, column=1).alignment = Alignment(horizontal="left")
    ws.merge_cells("A4:I4")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 70

    header_fill = PatternFill("solid", fgColor="DDE6F3")
    for col_idx, (name, width) in enumerate(_SEARCH_COLUMNS, start=1):
        cell = ws.cell(row=5, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = _BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = width

    for idx, rule in enumerate(_SAMPLE_SEARCH_RULES, start=6):
        for col, value in enumerate(rule, start=1):
            ws.cell(row=idx, column=col, value=value)

    ws.freeze_panes = "A6"

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
