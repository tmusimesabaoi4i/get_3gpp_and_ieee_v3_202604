"""Generate samples/sample_search.xlsx with the Sheet1 layout used by stdsearch.

Usage:
    python scripts/build_sample_search.py [--job <JobFolder path>] [--out <xlsx path>]

If --job is not provided, leave JobFolder empty (the user must fill it in).
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADER_FILL = PatternFill("solid", fgColor="DDE6F3")
HEADER_FONT = Font(bold=True)
NOTE_FILL = PatternFill("solid", fgColor="FFF7D6")

COLUMNS = [
    ("Use", 7),
    ("RuleName", 22),
    ("MatchType", 12),
    ("Term1", 22),
    ("Term2", 22),
    ("Term3", 18),
    ("Term4", 18),
    ("Scope", 14),
    ("Notes", 28),
]

SAMPLE_RULES = [
    # (Use, RuleName, MatchType, Term1, Term2, Term3, Term4, Scope, Notes)
    ("yes", "bssid_single",     "SINGLE", "Multiple BSSID", "",            "", "", "sentence",  "一語検索 (IEEE)"),
    ("yes", "bssid_sta_profile","AND",    "Multiple BSSID", "STA profile", "", "", "sentence",  "同一文にBSSIDとSTA profile"),
    ("yes", "mmwave_bssid",     "AND",    "mmWave",         "M-BSSID",     "", "", "block",     "スライド全体でAND (IEEE)"),
    ("yes", "dci_priority",     "AND",    "DCI format 0_3", "Priority indicator", "", "", "paragraph", "同一段落AND (3GPP Word)"),
    ("no",  "example_disabled", "SINGLE", "example",        "",            "", "", "sentence",  "Use=no で無効化例"),
]


def _paint_header(ws, row: int) -> None:
    for col, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = width


def build(out_path: Path, job_folder: str = "", output_folder: str = "") -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Top settings.
    ws["A1"] = "ProjectName"
    ws["B1"] = "bssid_search"
    ws["A2"] = "JobFolder"
    ws["B2"] = job_folder
    ws["A3"] = "OutputFolder"
    ws["B3"] = output_folder
    ws["A4"] = "(空欄可: JobFolder/search/YYYYMMDD_<ProjectName> に出力)"
    for row in (1, 2, 3):
        ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=4, column=1).fill = NOTE_FILL
    ws.cell(row=4, column=1).alignment = Alignment(horizontal="left")
    ws.merge_cells("A4:I4")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 70

    # Rule table header at row 5.
    _paint_header(ws, row=5)

    for idx, rule in enumerate(SAMPLE_RULES, start=6):
        for col, value in enumerate(rule, start=1):
            ws.cell(row=idx, column=col, value=value)

    ws.freeze_panes = "A6"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _parse_args() -> argparse.Namespace:
    here = Path(__file__).resolve().parent.parent
    default_out = here / "samples" / "sample_search.xlsx"
    p = argparse.ArgumentParser()
    p.add_argument("--out", type=Path, default=default_out,
                   help="Output xlsx path (default: stdsearch/samples/sample_search.xlsx)")
    p.add_argument("--job", type=str, default="",
                   help="JobFolder to pre-fill in B2 (optional).")
    p.add_argument("--output", type=str, default="",
                   help="OutputFolder to pre-fill in B3 (optional, empty = default).")
    return p.parse_args()


def main() -> int:
    args = _parse_args()
    out = build(args.out, job_folder=args.job, output_folder=args.output)
    print(f"Wrote: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
