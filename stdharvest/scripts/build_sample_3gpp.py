"""Generate samples/sample_download.xlsx for 3GPP testing.

Run:
    python scripts/build_sample_3gpp.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

HERE = Path(__file__).resolve().parent
SAMPLE_PATH = HERE.parent / "samples" / "sample_download.xlsx"

HEADER_FILL = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")
LABEL_FILL = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
BOLD = Font(bold=True)


def build() -> None:
    wb = Workbook()

    # ------------------------------------------------------------------ Sheet1
    ws1 = wb.active
    ws1.title = "Sheet1"

    ws1["A1"] = "SourceType"
    ws1["A1"].font = BOLD
    ws1["A1"].fill = LABEL_FILL
    ws1["B1"] = "3gpp"

    ws1["A2"] = "OutputRootFolder"
    ws1["A2"].font = BOLD
    ws1["A2"].fill = LABEL_FILL
    ws1["B2"] = r"C:\temp\std_docs"

    ws1["A3"] = "JobName"
    ws1["A3"].font = BOLD
    ws1["A3"].fill = LABEL_FILL
    ws1["B3"] = "3gpp_sample_job"

    # Row 4: header for the detail area.
    headers = ["", "", "Link", "Status", "SavedPath", "Message", "LastRunAt"]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws1.cell(row=4, column=col_idx, value=text)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")

    # 3GPP RAN1 documents from meeting #114b.
    base_url = "https://www.3gpp.org/ftp/TSG_RAN/WG1_RL1/TSGR1_114b/Docs"
    doc_ids = [
        "R1-2309136",
        "R1-2309171",
        "R1-2309306",
        "R1-2309685",
        "R1-2309698",
    ]
    samples = [(doc_id, f"{base_url}/{doc_id}.zip") for doc_id in doc_ids]

    for i, (title, url) in enumerate(samples, start=5):
        cell = ws1.cell(row=i, column=3, value=title)
        cell.hyperlink = url
        cell.font = Font(color="FF0563C1", underline="single")

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 40
    ws1.column_dimensions["C"].width = 48
    ws1.column_dimensions["D"].width = 14
    ws1.column_dimensions["E"].width = 40
    ws1.column_dimensions["F"].width = 40
    ws1.column_dimensions["G"].width = 22

    dv_source = DataValidation(type="list", formula1='"3gpp,ieee"', allow_blank=False)
    dv_source.error = "3gpp か ieee を選択してください"
    dv_source.errorTitle = "SourceType"
    ws1.add_data_validation(dv_source)
    dv_source.add("B1")

    # ------------------------------------------------------------------ Sheet2
    ws2 = wb.create_sheet("Sheet2")
    settings_rows = [
        ("ProxyURL", ""),
        ("TimeoutSec", 60),
        ("RetryCount", 3),
        ("SleepSec", 0.5),
        ("OverwriteExisting", "no"),
        ("MinFileSizeKB", 10),
        ("MaxFileSizeMB", 100),
        ("OnTooSmallFile", "error"),
        ("OnTooLargeFile", "skip"),
        ("KillOfficeAppsBeforeRun", "yes"),
        ("DownloadWorkers", 8),
        ("UnzipWorkers", 4),
        ("PdfWorkers", 2),
        ("HtmlWorkers", 6),
        ("CombineHtmlBatchSize", 5),
    ]
    for i, (label, value) in enumerate(settings_rows, start=1):
        c_a = ws2.cell(row=i, column=1, value=label)
        c_a.font = BOLD
        c_a.fill = LABEL_FILL
        ws2.cell(row=i, column=2, value=value)

    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 24

    validations = [
        ("B5", '"yes,no"'),
        ("B8", '"error,skip"'),
        ("B9", '"skip,pdf_only,keep_raw"'),
        ("B10", '"yes,no"'),
    ]
    for cell_ref, formula in validations:
        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
        ws2.add_data_validation(dv)
        dv.add(cell_ref)

    SAMPLE_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SAMPLE_PATH)
    print(f"Wrote {SAMPLE_PATH}")


if __name__ == "__main__":
    build()
