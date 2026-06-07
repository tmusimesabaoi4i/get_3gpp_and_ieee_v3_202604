"""Collect tab - GUI wrapper around stdharvest.

UX guarantees:

* Whenever you click "Open job folder" or "Open html/index.html", the path is
  re-resolved from the **current** Excel file (B1/B2/B3) - so editing the
  workbook between runs always opens the right place.
* Auto-detects when the chosen Excel changes on disk and reloads its preview
  (the "Reload" button forces a refresh too).
* Shows elapsed time + a per-tab state line while a job is running.
"""
from __future__ import annotations

import time
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Optional, Tuple

from .log_panel import LogPanel
from .osutil import open_in_shell
from .paths import app_log_dir
from .runner import JobMessage, JobRunner


def _today_compact() -> str:
    try:
        from stdharvest.utils import now_compact_date  # type: ignore
        return now_compact_date()
    except Exception:
        return datetime.now().strftime("%Y%m%d")


def _read_harvest_excel(path: Path) -> Tuple[str, str, str]:
    """Return (source_type, output_root, job_name) from a download workbook."""
    from openpyxl import load_workbook  # lazy
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
        source = (ws["B1"].value or "").strip().lower() if ws["B1"].value else ""
        out_root = (ws["B2"].value or "").strip() if ws["B2"].value else ""
        job_name = (ws["B3"].value or "").strip() if ws["B3"].value else ""
    finally:
        wb.close()
    return source, out_root, job_name


def _resolve_job_folder(out_root: str, job_name: str) -> Optional[Path]:
    """Pick the right job folder based on current Excel values.

    Strategy:
    1. ``<out_root>/<YYYYMMDD>_<job_name>`` for **today** if it exists.
    2. Otherwise the most recently modified ``<out_root>/*_<job_name>``.
    3. Otherwise None.
    """
    if not out_root or not job_name:
        return None
    root = Path(out_root)
    if not root.exists():
        return None
    today = root / f"{_today_compact()}_{job_name}"
    if today.exists():
        return today
    suffix = f"_{job_name}"
    candidates = sorted(
        (p for p in root.iterdir() if p.is_dir() and p.name.endswith(suffix)),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


class HarvestTab(ttk.Frame):
    """`stdharvest` GUI: pick Excel -> run -> open outputs."""

    def __init__(self, master: tk.Misc) -> None:
        super().__init__(master, padding=8)
        self._excel_path: Optional[Path] = None
        self._excel_mtime: Optional[float] = None

        self._runner = JobRunner(logger_names=("stdharvest",))
        self._job_started: Optional[float] = None

        self._build_ui()
        self._poll_job()
        self._poll_excel_mtime()

    # ------------------------------------------------------------------ UI
    def _build_ui(self) -> None:
        self.columnconfigure(1, weight=1)

        ttk.Label(self, text="Excel (sample_download.xlsx):").grid(
            row=0, column=0, sticky="w", padx=(0, 8), pady=4
        )
        self._excel_var = tk.StringVar()
        ttk.Entry(self, textvariable=self._excel_var, state="readonly").grid(
            row=0, column=1, sticky="ew", pady=4
        )
        btns = ttk.Frame(self)
        btns.grid(row=0, column=2, sticky="e", padx=(8, 0), pady=4)
        ttk.Button(btns, text="Browse...", command=self._browse_excel, width=10).pack(
            side=tk.LEFT
        )
        ttk.Button(btns, text="Reload", command=self._reload_excel, width=8).pack(
            side=tk.LEFT, padx=(4, 0)
        )

        info = ttk.LabelFrame(self, text="Job Preview", padding=6)
        info.grid(row=1, column=0, columnspan=3, sticky="ew", pady=6)
        info.columnconfigure(1, weight=1)

        labels = [
            ("SourceType:", "_src_var"),
            ("OutputRootFolder:", "_out_var"),
            ("JobName:", "_job_var"),
            ("JobFolder (today predicted):", "_jobfolder_var"),
            ("JobFolder (resolved):", "_resolved_var"),
        ]
        for r, (label, attr) in enumerate(labels):
            ttk.Label(info, text=label).grid(row=r, column=0, sticky="w", padx=(0, 6))
            var = tk.StringVar(value="-")
            setattr(self, attr, var)
            ttk.Label(info, textvariable=var, foreground="#222").grid(
                row=r, column=1, sticky="w"
            )

        self._preview_status = tk.StringVar(value="")
        ttk.Label(info, textvariable=self._preview_status, foreground="#a60").grid(
            row=len(labels), column=0, columnspan=2, sticky="w", pady=(4, 0)
        )

        actions = ttk.Frame(self)
        actions.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(0, 4))
        # Right-most column is the elastic gap so right-aligned widgets stick
        # to the right edge regardless of window width.
        actions.columnconfigure(3, weight=1)

        self._run_btn = ttk.Button(
            actions, text="Run harvest", command=self._on_run,
            state=tk.DISABLED, width=16,
        )
        self._run_btn.grid(row=0, column=0, sticky="w", padx=(0, 6), pady=2)
        self._open_folder_btn = ttk.Button(
            actions, text="Open job folder", command=self._open_job_folder,
            state=tk.DISABLED, width=18,
        )
        self._open_folder_btn.grid(row=0, column=1, sticky="w", padx=6, pady=2)
        self._open_index_btn = ttk.Button(
            actions, text="Open html/index.html", command=self._open_index,
            state=tk.DISABLED, width=22,
        )
        self._open_index_btn.grid(row=0, column=2, sticky="w", padx=6, pady=2)

        self._state_var = tk.StringVar(value="Idle")
        ttk.Label(actions, textvariable=self._state_var, foreground="#357").grid(
            row=1, column=0, columnspan=2, sticky="w", padx=(0, 6), pady=(4, 0),
        )
        self._progress = ttk.Progressbar(
            actions, mode="indeterminate", length=200,
        )
        self._progress.grid(row=1, column=3, sticky="e", pady=(4, 0))

        self._log = LogPanel(self, title="stdharvest log")
        self._log.grid(row=3, column=0, columnspan=3, sticky="nsew", pady=(8, 0))
        self.rowconfigure(3, weight=1)

    # ------------------------------------------------------------------ Excel
    def _browse_excel(self) -> None:
        initial = str(self._excel_path.parent) if self._excel_path else ""
        path = filedialog.askopenfilename(
            title="Select sample_download.xlsx",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
            initialdir=initial,
        )
        if not path:
            return
        self._set_excel(Path(path))

    def _reload_excel(self) -> None:
        if self._excel_path is None:
            messagebox.showinfo("No Excel", "Pick an Excel workbook first.")
            return
        self._refresh_preview(force=True)

    def _set_excel(self, path: Path) -> None:
        self._excel_path = path
        try:
            self._excel_mtime = path.stat().st_mtime
        except OSError:
            self._excel_mtime = None
        self._excel_var.set(str(path))
        self._refresh_preview(force=True)

    def _poll_excel_mtime(self) -> None:
        try:
            if self._excel_path is not None and self._excel_path.exists():
                mt = self._excel_path.stat().st_mtime
                if self._excel_mtime is not None and mt != self._excel_mtime:
                    self._excel_mtime = mt
                    self._log.append("Excel changed on disk - reloading preview.")
                    self._refresh_preview(force=False)
        finally:
            self.after(1500, self._poll_excel_mtime)

    def _refresh_preview(self, force: bool) -> None:
        if self._excel_path is None:
            return
        if not self._excel_path.exists():
            self._preview_status.set(f"Excel not found: {self._excel_path}")
            self._run_btn.configure(state=tk.DISABLED)
            return
        try:
            source, out_root, job_name = _read_harvest_excel(self._excel_path)
        except Exception as exc:  # noqa: BLE001
            self._preview_status.set(f"Failed to read Excel: {exc}")
            for var in (self._src_var, self._out_var, self._job_var,
                        self._jobfolder_var, self._resolved_var):
                var.set("-")
            self._run_btn.configure(state=tk.DISABLED)
            return

        job_name = job_name or "stdharvest_job"
        self._src_var.set(source or "(empty)")
        self._out_var.set(out_root or "(empty)")
        self._job_var.set(job_name)

        if out_root:
            predicted = Path(out_root) / f"{_today_compact()}_{job_name}"
            self._jobfolder_var.set(str(predicted))
        else:
            self._jobfolder_var.set("-")

        resolved = _resolve_job_folder(out_root, job_name) if out_root else None
        if resolved is not None:
            self._resolved_var.set(str(resolved))
            self._open_folder_btn.configure(state=tk.NORMAL)
            idx = resolved / "html" / "index.html"
            self._open_index_btn.configure(
                state=tk.NORMAL if idx.exists() else tk.DISABLED
            )
        else:
            self._resolved_var.set("(none yet - run a job to create)")
            self._open_folder_btn.configure(state=tk.DISABLED)
            self._open_index_btn.configure(state=tk.DISABLED)

        problems = []
        if source not in ("3gpp", "ieee"):
            problems.append("SourceType must be '3gpp' or 'ieee'")
        if not out_root:
            problems.append("OutputRootFolder is empty")
        if problems:
            self._preview_status.set(" / ".join(problems))
            self._run_btn.configure(state=tk.DISABLED)
        else:
            self._preview_status.set("")
            if not self._runner.is_running:
                self._run_btn.configure(state=tk.NORMAL)
        if force:
            self._log.append(f"Loaded Excel: {self._excel_path}")

    # ------------------------------------------------------------------ run
    def _on_run(self) -> None:
        if self._runner.is_running:
            messagebox.showinfo("Busy", "A harvest job is already running.")
            return
        if self._excel_path is None or not self._excel_path.exists():
            messagebox.showwarning("Excel required", "Select sample_download.xlsx first.")
            return
        try:
            from stdharvest.cli import run as harvest_run  # type: ignore
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror(
                "stdharvest not installed",
                f"Failed to import stdharvest: {exc}\n\n"
                "Install it with: pip install -e ./stdharvest",
            )
            return

        excel = self._excel_path.resolve()
        self._log.append(f"=== Starting harvest: {excel} ===")
        self._run_btn.configure(state=tk.DISABLED)
        self._open_folder_btn.configure(state=tk.DISABLED)
        self._open_index_btn.configure(state=tk.DISABLED)
        self._progress.configure(mode="indeterminate")
        self._progress.start(80)
        self._job_started = time.time()
        self._tick_state(running=True)
        self._runner.start(
            target=lambda: harvest_run(excel),
            log_file_dir=app_log_dir() / "harvest",
            log_file_prefix="harvest",
        )

    def _tick_state(self, running: bool) -> None:
        if running and self._runner.is_running and self._job_started is not None:
            elapsed = int(time.time() - self._job_started)
            self._state_var.set(
                f"Running... {elapsed // 60:02d}:{elapsed % 60:02d}"
            )
            self.after(500, lambda: self._tick_state(True))
        elif not running:
            return

    def _poll_job(self) -> None:
        try:
            for msg in self._runner.drain():
                self._handle_message(msg)
        finally:
            self.after(120, self._poll_job)

    def _handle_message(self, msg: JobMessage) -> None:
        if msg.kind == "log":
            self._log.append(msg.text)
            return
        if msg.kind in ("done", "error"):
            self._progress.stop()
            self._progress.configure(mode="determinate")
            self._progress["value"] = 0
            elapsed = (
                int(time.time() - self._job_started) if self._job_started else 0
            )
            self._state_var.set(
                f"{'Done' if msg.kind == 'done' else 'Failed'} "
                f"({elapsed // 60:02d}:{elapsed % 60:02d})"
            )
            self._job_started = None
            self._log.append(f"=== {msg.text} ===")
            # Re-resolve preview so Open buttons reflect post-run reality.
            self._refresh_preview(force=False)

    # ------------------------------------------------------------------ open
    def _open_job_folder(self) -> None:
        path = self._latest_job_folder(report_errors=True)
        if path is None:
            return
        try:
            open_in_shell(path)
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Error", f"Failed to open folder: {exc}")

    def _open_index(self) -> None:
        path = self._latest_job_folder(report_errors=True)
        if path is None:
            return
        idx = path / "html" / "index.html"
        if not idx.exists():
            messagebox.showwarning(
                "index.html missing",
                f"Cannot find {idx}\n\nRun the harvest job first.",
            )
            return
        try:
            open_in_shell(idx)
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Error", f"Failed to open index.html: {exc}")

    def _latest_job_folder(self, *, report_errors: bool) -> Optional[Path]:
        """Always re-read the Excel and re-resolve the job folder."""
        if self._excel_path is None or not self._excel_path.exists():
            if report_errors:
                messagebox.showwarning("No Excel", "Pick an Excel workbook first.")
            return None
        try:
            _src, out_root, job_name = _read_harvest_excel(self._excel_path)
        except Exception as exc:  # noqa: BLE001
            if report_errors:
                messagebox.showerror(
                    "Excel read failed",
                    f"Failed to read {self._excel_path}: {exc}",
                )
            return None
        job_name = job_name or "stdharvest_job"
        resolved = _resolve_job_folder(out_root, job_name)
        if resolved is None and report_errors:
            messagebox.showwarning(
                "Job folder missing",
                f"Cannot find any folder matching '*_{job_name}' under "
                f"'{out_root}'. Run the harvest job first.",
            )
        return resolved
