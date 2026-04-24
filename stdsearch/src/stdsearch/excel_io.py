"""Read Search.xlsx (Sheet1 only)."""
from __future__ import annotations

import logging
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook

from .models import (
    MATCH_AND,
    MATCH_SINGLE,
    SCOPE_BLOCK,
    SCOPE_PARAGRAPH,
    SCOPE_SECTION,
    SCOPE_SENTENCE,
    SearchProject,
    SearchRule,
    VALID_MATCH_TYPES,
    VALID_SCOPES,
)

logger = logging.getLogger(__name__)

SHEET_NAME = "Sheet1"
FIRST_RULE_ROW = 5

# Column positions (1-based) for the rule table.
COL_USE = 1          # A
COL_RULE = 2         # B
COL_MATCH = 3        # C
COL_TERM1 = 4        # D
COL_TERM2 = 5        # E
COL_TERM3 = 6        # F
COL_TERM4 = 7        # G
COL_SCOPE = 8        # H
COL_NOTES = 9        # I


def _str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_use(value) -> bool:
    token = _str(value).lower()
    return token in ("yes", "y", "1", "true", "on")


def _parse_match_type(value) -> str:
    token = _str(value).upper()
    if token in VALID_MATCH_TYPES:
        return token
    return MATCH_SINGLE


def _parse_scope(value) -> str:
    token = _str(value).lower()
    if token in VALID_SCOPES:
        return token
    return SCOPE_SENTENCE


def read_search_excel(excel_path: Path) -> SearchProject:
    if not excel_path.exists():
        raise FileNotFoundError(excel_path)
    wb = load_workbook(excel_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found in {excel_path}")
    ws = wb[SHEET_NAME]

    project_name = _str(ws["B1"].value) or "unnamed"
    job_folder_str = _str(ws["B2"].value)
    if not job_folder_str:
        raise RuntimeError("Sheet1!B2 (JobFolder) is empty")
    job_folder = Path(job_folder_str)
    output_str = _str(ws["B3"].value)
    output_folder: Optional[Path] = Path(output_str) if output_str else None

    rules: List[SearchRule] = []
    for row_no in range(FIRST_RULE_ROW, ws.max_row + 1):
        use_val = ws.cell(row=row_no, column=COL_USE).value
        rule_name = _str(ws.cell(row=row_no, column=COL_RULE).value)
        if not rule_name and use_val is None:
            continue
        terms = [
            _str(ws.cell(row=row_no, column=c).value)
            for c in (COL_TERM1, COL_TERM2, COL_TERM3, COL_TERM4)
        ]
        terms = [t for t in terms if t]
        rule = SearchRule(
            row_no=row_no,
            use=_parse_use(use_val),
            rule_name=rule_name or f"rule_{row_no}",
            match_type=_parse_match_type(ws.cell(row=row_no, column=COL_MATCH).value),
            terms=terms,
            scope=_parse_scope(ws.cell(row=row_no, column=COL_SCOPE).value),
            notes=_str(ws.cell(row=row_no, column=COL_NOTES).value),
        )
        if not rule.terms:
            logger.warning("Rule %s (row %d) has no search terms; skipped.", rule.rule_name, row_no)
            continue
        if rule.match_type == MATCH_AND and len(rule.terms) < 2:
            logger.warning(
                "Rule %s (row %d) is AND but has only %d term(s); it will behave like SINGLE.",
                rule.rule_name, row_no, len(rule.terms),
            )
            rule.match_type = MATCH_SINGLE
        rules.append(rule)

    return SearchProject(
        project_name=project_name,
        job_folder=job_folder,
        output_folder=output_folder,
        rules=rules,
    )
