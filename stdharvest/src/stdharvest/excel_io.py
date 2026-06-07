"""Read Sheet1 / Sheet2 and write back results on Sheet1 columns D-G.

All writes are batched into a single save() call from the caller.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import List, Tuple

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from . import models
from .models import DownloadRow, JobContext, Settings, VALID_SOURCES
from .utils import now_compact_date, now_iso

logger = logging.getLogger(__name__)

SHEET_DOWNLOAD = "Sheet1"
SHEET_SETTINGS = "Sheet2"

FIRST_DATA_ROW = 5  # Sheet1 data starts at row 5


@dataclass
class ExcelData:
    workbook: Workbook
    job: JobContext
    settings: Settings
    rows: List[DownloadRow]        # rows that will actually be processed
    skipped_rows: List[DownloadRow]  # kept to preserve their existing cells


def _str(cell_value) -> str:
    if cell_value is None:
        return ""
    return str(cell_value).strip()


def _bool_yes_no(value, default: bool) -> bool:
    token = _str(value).lower()
    if token == "yes":
        return True
    if token == "no":
        return False
    return default


def _int(value, default: int) -> int:
    try:
        return int(float(_str(value)))
    except (TypeError, ValueError):
        return default


def _float(value, default: float) -> float:
    try:
        return float(_str(value))
    except (TypeError, ValueError):
        return default


def _read_settings(wb: Workbook) -> Settings:
    if SHEET_SETTINGS not in wb.sheetnames:
        logger.warning("Sheet2 (Settings) not found. Using defaults.")
        return Settings()
    ws = wb[SHEET_SETTINGS]

    def v(row: int):
        return ws.cell(row=row, column=2).value

    s = Settings()
    s.proxy_url = _str(v(1))
    s.timeout_sec = _int(v(2), s.timeout_sec)
    s.retry_count = _int(v(3), s.retry_count)
    s.sleep_sec = _float(v(4), s.sleep_sec)
    s.overwrite_existing = _bool_yes_no(v(5), s.overwrite_existing)
    s.min_file_size_kb = _int(v(6), s.min_file_size_kb)
    s.max_file_size_mb = _int(v(7), s.max_file_size_mb)
    s.on_too_small_file = _str(v(8)).lower() or s.on_too_small_file
    s.on_too_large_file = _str(v(9)).lower() or s.on_too_large_file
    s.kill_office_apps_before_run = _bool_yes_no(v(10), s.kill_office_apps_before_run)
    s.download_workers = _int(v(11), s.download_workers)
    s.unzip_workers = _int(v(12), s.unzip_workers)
    s.pdf_workers = _int(v(13), s.pdf_workers)
    s.html_workers = _int(v(14), s.html_workers)
    s.combine_html_batch_size = _int(v(15), s.combine_html_batch_size)
    # Optional proxy authentication (appended rows; absent in older workbooks).
    s.proxy_user = _str(v(16))
    s.proxy_password = _str(v(17))
    return s


def _read_job(wb: Workbook, excel_path: Path, run_started_at: str) -> JobContext:
    if SHEET_DOWNLOAD not in wb.sheetnames:
        raise RuntimeError(f"Sheet1 '{SHEET_DOWNLOAD}' not found in {excel_path}")
    ws = wb[SHEET_DOWNLOAD]
    source = _str(ws["B1"].value).lower()
    if source not in VALID_SOURCES:
        raise RuntimeError(
            f"Sheet1!B1 must be one of {sorted(VALID_SOURCES)} (got '{source}')"
        )
    output_root = _str(ws["B2"].value)
    if not output_root:
        raise RuntimeError("Sheet1!B2 (OutputRootFolder) is empty")
    job_name = _str(ws["B3"].value) or "stdharvest_job"

    job_folder = Path(output_root) / f"{now_compact_date()}_{job_name}"
    return JobContext(
        source_type=source,
        output_root=Path(output_root),
        job_name=job_name,
        job_folder=job_folder,
        run_started_at=run_started_at,
    )


def _extract_hyperlink_url(cell) -> str:
    hl = getattr(cell, "hyperlink", None)
    if hl is not None:
        target = getattr(hl, "target", None)
        if target:
            return str(target).strip()
    # Also accept plain URL text if no hyperlink was attached.
    val = _str(cell.value)
    if val.lower().startswith(("http://", "https://", "ftp://")):
        return val
    return ""


def read_excel(excel_path: Path) -> ExcelData:
    """Read the workbook and return job / settings / rows.

    Rows whose Status is DONE / SKIPPED / DONE_WITH_SKIP are preserved but
    not returned as processable rows (they'll be skipped).
    """
    if not excel_path.exists():
        raise FileNotFoundError(excel_path)
    wb = load_workbook(excel_path, data_only=False)
    run_started_at = now_iso()
    job = _read_job(wb, excel_path, run_started_at)
    settings = _read_settings(wb)

    ws = wb[SHEET_DOWNLOAD]
    process: List[DownloadRow] = []
    skipped: List[DownloadRow] = []

    seq = 0
    for row_no in range(FIRST_DATA_ROW, ws.max_row + 1):
        link_cell = ws.cell(row=row_no, column=3)  # C
        title = _str(link_cell.value)
        url = _extract_hyperlink_url(link_cell)
        existing_status = _str(ws.cell(row=row_no, column=4).value).upper()
        if not title and not url:
            # Empty line, ignore silently.
            continue

        if existing_status in models.RETRY_STATUSES:
            seq += 1
            process.append(
                DownloadRow(
                    row_no=row_no,
                    seq=seq,
                    title=title or url,
                    url=url,
                    existing_status=existing_status,
                )
            )
        else:
            # DONE / SKIPPED / DONE_WITH_SKIP / other -> keep existing values.
            skipped.append(
                DownloadRow(
                    row_no=row_no,
                    seq=0,
                    title=title or url,
                    url=url,
                    existing_status=existing_status,
                    status=existing_status,
                    saved_path=_str(ws.cell(row=row_no, column=5).value),
                    message=_str(ws.cell(row=row_no, column=6).value),
                    last_run_at=_str(ws.cell(row=row_no, column=7).value),
                )
            )

    return ExcelData(
        workbook=wb,
        job=job,
        settings=settings,
        rows=process,
        skipped_rows=skipped,
    )


def write_back(excel_path: Path, data: ExcelData) -> None:
    """Batch-write D-G columns for processed rows and save the workbook."""
    wb = data.workbook
    ws = wb[SHEET_DOWNLOAD]
    for row in data.rows:
        ws.cell(row=row.row_no, column=4, value=row.status or "")
        ws.cell(row=row.row_no, column=5, value=row.saved_path or "")
        ws.cell(row=row.row_no, column=6, value=row.message or "")
        ws.cell(row=row.row_no, column=7, value=row.last_run_at or "")
    wb.save(excel_path)


def total_row_count(data: ExcelData) -> Tuple[int, int]:
    """Return (processed_count, skipped_count) for reporting."""
    return len(data.rows), len(data.skipped_rows)
